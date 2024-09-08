import xml.etree.ElementTree as ET
import os
import csv
import openpyxl  # Biblioteca para manejar Excel
import tkinter as tk
from tkinter import filedialog, scrolledtext
import threading
import re  # Biblioteca para manejar expresiones regulares

# Diccionario que contiene variantes incorrectas y correcciones para cada provincia
def normalizar_provincia(provincia):
    provincia_variantes = {
        "Córdoba": ["CÓRODOBA", "CC3rdoba", "Cordoba", "C rdoba", "C rdob"],
        "Araba/Álava": ["Alava", "Araba", "Araba Alava"],
        "Albacete": ["Alvacete", "Alvace"],
        "Alicante": ["Alacant", "Alicate"],
        "Almería": ["Almeria", "Almerìa"],
        "Ávila": ["Avila", "Avìla"],
        "Badajoz": ["Badagoj", "Badajó"],
        "Illes Balears": ["Baleares", "Illes", "Islas Baleares"],
        "Barcelona": ["Barna", "Barcelo", "Barselona"],
        "Burgos": ["Burgo", "Burgues"],
        "Cáceres": ["Caceres", "Càceres"],
        "Cádiz": ["Cadiz", "Cadìz"],
        "Castellón": ["Castellon", "Castellò"],
        "Ciudad Real": ["C.Real", "Ciudad R."],
        "Coruña": ["La Coruña", "A Coruña"],
        "Cuenca": ["Cuenca"],
        "Girona": ["Gerona"],
        "Granada": ["Granà", "Grnada"],
        "Guadalajara": ["Guada"],
        "Gipuzkoa": ["Guipúzcoa", "Guipozkoa"],
        "Huelva": ["Huelba"],
        "Huesca": ["Huesk"],
        "Jaén": ["Jaen", "Jèn"],
        "León": ["Leon", "Léon"],
        "Lleida": ["Lerida"],
        "La Rioja": ["Rioja"],
        "Lugo": ["Lugo"],
        "Madrid": ["Madrì", "Madird"],
        "Málaga": ["Malaga", "Màlaga"],
        "Murcia": ["Murcìa", "Murçia"],
        "Navarra": ["Nafarroa"],
        "Ourense": ["Orense"],
        "Asturias": ["Asturies", "Asturìa"],
        "Palencia": ["Palencia"],
        "Las Palmas": ["Palmas", "Gran Canaria"],
        "Pontevedra": ["Ponte"],
        "Salamanca": ["Salmanaca"],
        "S.C. Tenerife": ["Tenerife", "Santa Cruz"],
        "Cantabria": ["Santander"],
        "Segovia": ["Segòvia"],
        "Sevilla": ["Sevilia", "Sevìlla"],
        "Soria": ["Sòria"],
        "Tarragona": ["Tarragòna", "Tarrag"],
        "Teruel": ["Terùel"],
        "Toledo": ["Toled"],
        "Valencia": ["València", "Valençia"],
        "Valladolid": ["Vallad", "Vallado"],
        "Bizkaia": ["Vizcaya", "Biskaia"],
        "Zamora": ["Zamòra"],
        "Zaragoza": ["Saragosa"],
        "Ceuta": ["Ceuat"],
        "Melilla": ["Melila", "Melill"]
    }
    
   
    # Diccionario para códigos postales según las provincias
    codigo_postal_provincia = {
        "01": "Araba/Álava", "02": "Albacete", "03": "Alicante", "04": "Almería", "05": "Ávila", 
        "06": "Badajoz", "07": "Illes Balears", "08": "Barcelona", "09": "Burgos", "10": "Cáceres", 
        "11": "Cádiz", "12": "Castellón", "13": "Ciudad Real", "14": "Córdoba", "15": "Coruña", 
        "16": "Cuenca", "17": "Girona", "18": "Granada", "19": "Guadalajara", "20": "Gipuzkoa", 
        "21": "Huelva", "22": "Huesca", "23": "Jaén", "24": "León", "25": "Lleida", "26": "La Rioja", 
        "27": "Lugo", "28": "Madrid", "29": "Málaga", "30": "Murcia", "31": "Navarra", "32": "Ourense", 
        "33": "Asturias", "34": "Palencia", "35": "Las Palmas", "36": "Pontevedra", "37": "Salamanca", 
        "38": "S.C. Tenerife", "39": "Cantabria", "40": "Segovia", "41": "Sevilla", "42": "Soria", 
        "43": "Tarragona", "44": "Teruel", "45": "Toledo", "46": "Valencia", "47": "Valladolid", 
        "48": "Bizkaia", "49": "Zamora", "50": "Zaragoza", "51": "Ceuta", "52": "Melilla"
    }
    
    # Verificar si la provincia tiene errores tipográficos y necesita corrección
    for provincia_correcta, variantes in provincia_variantes.items():
        if provincia in variantes:
            return provincia_correcta  # Devolver la provincia corregida
    
    # Si no está mal escrita, devolver la provincia original
    return provincia

# Función para guardar un registro en CSV
def guardar_registro_csv(writer, registro, consola, archivo_csv):
    try:
        writer.writerow(registro)
        consola.insert(tk.END, f"Guardado en {archivo_csv}: {registro['Referencia Catastral']}\n")
        consola.see(tk.END)
    except Exception as e:
        consola.insert(tk.END, f"Error al guardar el registro en {archivo_csv}: {e}\n")
        consola.see(tk.END)

# Función para guardar un registro en Excel
def guardar_registro_excel(ws, registro, row, consola, archivo_excel):
    try:
        for col, key in enumerate(registro.keys(), 1):
            ws.cell(row=row, column=col, value=registro[key])
        consola.insert(tk.END, f"Guardado en {archivo_excel}: {registro['Referencia Catastral']}\n")
        consola.see(tk.END)
    except Exception as e:
        consola.insert(tk.END, f"Error al guardar el registro en {archivo_excel}: {e}\n")
        consola.see(tk.END)

# Función para limpiar y normalizar el código postal (eliminar puntos)
def normalizar_codigo_postal(codigo_postal):
    return codigo_postal.replace('.', '').replace(' ', '')

# Función para guardar incidencias en CSV y Excel
def guardar_incidencia_csv_excel(writer_incidencias_csv, ws_incidencias, registro, consola, archivo_csv_incidencias, archivo_excel_incidencias, row_incidencias):
    try:
        writer_incidencias_csv.writerow(registro)
        for col, key in enumerate(registro.keys(), 1):
            ws_incidencias.cell(row=row_incidencias, column=col, value=registro[key])
        consola.insert(tk.END, f"Incidencia guardada en {archivo_csv_incidencias}: {registro['Referencia Catastral']}\n")
        consola.see(tk.END)
    except Exception as e:
        consola.insert(tk.END, f"Error al guardar la incidencia en {archivo_csv_incidencias}: {e}\n")
        consola.see(tk.END)

# Función para validar si el código postal coincide con la provincia
def validar_codigo_postal_provincia(provincia_normalizada, codigo_postal_normalizado):
    # Diccionario que asocia los primeros dos dígitos del código postal con la provincia
    codigo_postal_provincia = {
        "01": "Araba/Álava", "02": "Albacete", "03": "Alicante", "04": "Almería", "05": "Ávila", 
        "06": "Badajoz", "07": "Illes Balears", "08": "Barcelona", "09": "Burgos", "10": "Cáceres", 
        "11": "Cádiz", "12": "Castellón", "13": "Ciudad Real", "14": "Córdoba", "15": "Coruña", 
        "16": "Cuenca", "17": "Girona", "18": "Granada", "19": "Guadalajara", "20": "Gipuzkoa", 
        "21": "Huelva", "22": "Huesca", "23": "Jaén", "24": "León", "25": "Lleida", "26": "La Rioja", 
        "27": "Lugo", "28": "Madrid", "29": "Málaga", "30": "Murcia", "31": "Navarra", "32": "Ourense", 
        "33": "Asturias", "34": "Palencia", "35": "Las Palmas", "36": "Pontevedra", "37": "Salamanca", 
        "38": "S.C. Tenerife", "39": "Cantabria", "40": "Segovia", "41": "Sevilla", "42": "Soria", 
        "43": "Tarragona", "44": "Teruel", "45": "Toledo", "46": "Valencia", "47": "Valladolid", 
        "48": "Bizkaia", "49": "Zamora", "50": "Zaragoza", "51": "Ceuta", "52": "Melilla"
    }
    
    # Obtenemos los primeros dos dígitos del código postal
    codigo_provincia = codigo_postal_normalizado[:2]
    
    # Validamos si los primeros dos dígitos del código postal coinciden con la provincia esperada
    if codigo_provincia in codigo_postal_provincia:
        return codigo_postal_provincia[codigo_provincia] == provincia_normalizada
    return False

# Función principal para procesar los registros
def indexar_archivos_xml_y_guardar(archivos_xml, archivo_csv_comas_base, archivo_csv_tabs_base, archivo_excel_base, consola, limite_registros=30000):
    total_registros = 0
    total_incidencias = 0
    num_archivo_bueno = 1
    num_archivo_incidencias = 1
    registros_buenos_por_archivo = 0
    registros_incidencias_por_archivo = 0
    row_incidencias = 1
    
    fieldnames = [
        'Referencia Catastral', 'Dirección', 'Provincia', 'Municipio', 
        'Código Postal', 'Año Construcción', 'Superficie Habitable', 'Demanda ACS', 
        'Emisiones CO2 Global', 'Calificación Global', 'Fecha Certificación'
    ]
    
    # Función para generar nombres de archivos
    def generar_nombres_archivos_buenos(num_archivo):
        archivo_csv_comas = archivo_csv_comas_base.replace('.csv', f'_comas_buenos_{num_archivo}.csv')
        archivo_csv_tabs = archivo_csv_tabs_base.replace('.csv', f'_tabulaciones_buenos_{num_archivo}.csv')
        archivo_excel = archivo_excel_base.replace('.xlsx', f'_buenos_{num_archivo}.xlsx')
        return archivo_csv_comas, archivo_csv_tabs, archivo_excel
    
    def generar_nombres_archivos_incidencias(num_archivo):
        archivo_csv_incidencias = archivo_csv_comas_base.replace('.csv', f'_incidencias_{num_archivo}.csv')
        archivo_excel_incidencias = archivo_excel_base.replace('.xlsx', f'_incidencias_{num_archivo}.xlsx')
        return archivo_csv_incidencias, archivo_excel_incidencias

    # Inicializamos los primeros archivos
    archivo_csv_comas, archivo_csv_tabs, archivo_excel_buenos = generar_nombres_archivos_buenos(num_archivo_bueno)
    archivo_csv_incidencias, archivo_excel_incidencias = generar_nombres_archivos_incidencias(num_archivo_incidencias)

    try:
        with open(archivo_csv_comas, 'w', newline='', encoding='utf-8') as f_comas, \
             open(archivo_csv_tabs, 'w', newline='', encoding='utf-8') as f_tabs, \
             open(archivo_csv_incidencias, 'w', newline='', encoding='utf-8') as f_incidencias:
            
            writer_comas = csv.DictWriter(f_comas, fieldnames=fieldnames, delimiter=',')
            writer_tabs = csv.DictWriter(f_tabs, fieldnames=fieldnames, delimiter='\t')
            writer_incidencias_csv = csv.DictWriter(f_incidencias, fieldnames=fieldnames, delimiter=',')
            
            writer_comas.writeheader()
            writer_tabs.writeheader()
            writer_incidencias_csv.writeheader()

            wb_buenos = openpyxl.Workbook()
            ws_buenos = wb_buenos.active
            ws_buenos.append(fieldnames)
            
            wb_incidencias = openpyxl.Workbook()
            ws_incidencias = wb_incidencias.active
            ws_incidencias.append(fieldnames)

            for archivo in archivos_xml:
                consola.insert(tk.END, f"Procesando archivo: {archivo}\n")
                consola.update_idletasks()

                try:
                    for event, elem in ET.iterparse(archivo, events=("end",)):
                        if elem.tag == "DatosEnergeticosDelEdificio":
                            provincia = elem.find(".//Provincia").text or "N/A"
                            codigo_postal = elem.find(".//CodigoPostal").text or "N/A"
                            
                            # Normalizar provincia y código postal
                            provincia_normalizada = normalizar_provincia(provincia, codigo_postal)
                            codigo_postal_normalizado = normalizar_codigo_postal(codigo_postal)

                            registro = {
                                'Referencia Catastral': elem.find(".//ReferenciaCatastral").text or "N/A",
                                'Dirección': elem.find(".//Direccion").text or "N/A",
                                'Provincia': provincia_normalizada,
                                'Municipio': elem.find(".//Municipio").text or "N/A",
                                'Código Postal': codigo_postal_normalizado,
                                'Año Construcción': elem.find(".//AnoConstruccion").text or "N/A",
                                'Superficie Habitable': elem.find(".//SuperficieHabitable").text or "N/A",
                                'Demanda ACS': elem.find(".//DemandaDiariaACS").text or "N/A",
                                'Emisiones CO2 Global': elem.find(".//EmisionesCO2/Global").text or "N/A",
                                'Calificación Global': elem.find(".//Calificacion/EnergiaPrimariaNoRenovable/Global").text or "N/A",
                                'Fecha Certificación': elem.find(".//DatosDelCertificador/Fecha").text or "N/A",
                            }

                            # Validar si el código postal coincide con la provincia
                            if not validar_codigo_postal_provincia(provincia_normalizada, codigo_postal_normalizado):
                                # Si el código postal no coincide con la provincia, es una incidencia
                                guardar_incidencia_csv_excel(writer_incidencias_csv, ws_incidencias, registro, consola, archivo_csv_incidencias, archivo_excel_incidencias, row_incidencias)
                                row_incidencias += 1
                                total_incidencias += 1
                                registros_incidencias_por_archivo += 1

                                # Si se alcanzan los 30,000 registros de incidencias
                                if registros_incidencias_por_archivo >= limite_registros:
                                    consola.insert(tk.END, f"Guardando archivo de incidencias {num_archivo_incidencias}...\n")
                                    wb_incidencias.save(archivo_excel_incidencias)
                                    num_archivo_incidencias += 1
                                    archivo_csv_incidencias, archivo_excel_incidencias = generar_nombres_archivos_incidencias(num_archivo_incidencias)
                                    wb_incidencias = openpyxl.Workbook()
                                    ws_incidencias = wb_incidencias.active
                                    ws_incidencias.append(fieldnames)
                                    registros_incidencias_por_archivo = 0
                            else:
                                # Guardar como registro válido
                                guardar_registro_csv(writer_comas, registro, consola, archivo_csv_comas)
                                guardar_registro_csv(writer_tabs, registro, consola, archivo_csv_tabs)
                                guardar_registro_excel(ws_buenos, registro, registros_buenos_por_archivo + 1, consola, archivo_excel_buenos)
                                total_registros += 1
                                registros_buenos_por_archivo += 1

                                # Si se alcanzan los 30,000 registros buenos
                                if registros_buenos_por_archivo >= limite_registros:
                                    consola.insert(tk.END, f"Guardando archivo de registros buenos {num_archivo_bueno}...\n")
                                    wb_buenos.save(archivo_excel_buenos)
                                    num_archivo_bueno += 1
                                    archivo_csv_comas, archivo_csv_tabs, archivo_excel_buenos = generar_nombres_archivos_buenos(num_archivo_bueno)
                                    wb_buenos = openpyxl.Workbook()
                                    ws_buenos = wb_buenos.active
                                    ws_buenos.append(fieldnames)
                                    registros_buenos_por_archivo = 0

                            elem.clear()

                    consola.insert(tk.END, f"Total registros procesados en {archivo}: {total_registros}\n")
                    consola.see(tk.END)

                except Exception as e:
                    consola.insert(tk.END, f"Error procesando {archivo}: {e}\n")
                    consola.see(tk.END)

            # Guardar los archivos finales si no alcanzaron el límite de 30,000 registros
            wb_buenos.save(archivo_excel_buenos)
            wb_incidencias.save(archivo_excel_incidencias)
            consola.insert(tk.END, f"Archivo Excel de registros buenos guardado en: {archivo_excel_buenos}\n")
            consola.insert(tk.END, f"Archivo Excel de incidencias guardado en: {archivo_excel_incidencias}\n")

    except Exception as e:
        consola.insert(tk.END, f"Error al abrir los archivos CSV: {e}\n")
        consola.see(tk.END)

    consola.insert(tk.END, f"Indexación completada. Total de registros buenos: {total_registros}, total de incidencias: {total_incidencias}\n")
    consola.see(tk.END)
# Diccionario de provincias y sus códigos postales
codigo_postal_provincia = {
    "01": "Araba/Álava", "02": "Albacete", "03": "Alicante", "04": "Almería", "05": "Ávila", 
    "06": "Badajoz", "07": "Illes Balears", "08": "Barcelona", "09": "Burgos", "10": "Cáceres", 
    "11": "Cádiz", "12": "Castellón", "13": "Ciudad Real", "14": "Córdoba", "15": "Coruña", 
    "16": "Cuenca", "17": "Girona", "18": "Granada", "19": "Guadalajara", "20": "Gipuzkoa", 
    "21": "Huelva", "22": "Huesca", "23": "Jaén", "24": "León", "25": "Lleida", "26": "La Rioja", 
    "27": "Lugo", "28": "Madrid", "29": "Málaga", "30": "Murcia", "31": "Navarra", "32": "Ourense", 
    "33": "Asturias", "34": "Palencia", "35": "Las Palmas", "36": "Pontevedra", "37": "Salamanca", 
    "38": "S.C. Tenerife", "39": "Cantabria", "40": "Segovia", "41": "Sevilla", "42": "Soria", 
    "43": "Tarragona", "44": "Teruel", "45": "Toledo", "46": "Valencia", "47": "Valladolid", 
    "48": "Bizkaia", "49": "Zamora", "50": "Zaragoza", "51": "Ceuta", "52": "Melilla"
}

# Función para validar si el código postal coincide con la provincia
def validar_codigo_postal_provincia(provincia_normalizada, codigo_postal_normalizado):
    # Obtenemos los primeros dos dígitos del código postal
    codigo_provincia = codigo_postal_normalizado[:2]
    
    # Validamos si los primeros dos dígitos del código postal coinciden con la provincia esperada
    if codigo_provincia in codigo_postal_provincia:
        return codigo_postal_provincia[codigo_provincia] == provincia_normalizada
    return False

# Función para procesar los registros del XML y separarlos en buenos e incidencias
def indexar_archivos_xml_y_guardar(archivos_xml, archivo_csv_comas_base, archivo_csv_tabs_base, archivo_excel_base, consola, limite_registros=30000):
    total_registros = 0
    total_incidencias = 0
    num_archivo_bueno = 1
    num_archivo_incidencias = 1
    registros_buenos_por_archivo = 0
    registros_incidencias_por_archivo = 0
    row_incidencias = 1
    
    fieldnames = [
        'Referencia Catastral', 'Dirección', 'Provincia', 'Municipio', 
        'Código Postal', 'Año Construcción', 'Superficie Habitable', 'Demanda ACS', 
        'Emisiones CO2 Global', 'Calificación Global', 'Fecha Certificación'
    ]
    
    # Función para generar nombres de archivos buenos
    def generar_nombres_archivos_buenos(num_archivo):
        archivo_csv_comas = archivo_csv_comas_base.replace('.csv', f'_comas_buenos_{num_archivo}.csv')
        archivo_csv_tabs = archivo_csv_tabs_base.replace('.csv', f'_tabulaciones_buenos_{num_archivo}.csv')
        archivo_excel = archivo_excel_base.replace('.xlsx', f'_buenos_{num_archivo}.xlsx')
        return archivo_csv_comas, archivo_csv_tabs, archivo_excel
    
    # Función para generar nombres de archivos de incidencias
    def generar_nombres_archivos_incidencias(num_archivo):
        archivo_csv_incidencias = archivo_csv_comas_base.replace('.csv', f'_incidencias_{num_archivo}.csv')
        archivo_excel_incidencias = archivo_excel_base.replace('.xlsx', f'_incidencias_{num_archivo}.xlsx')
        return archivo_csv_incidencias, archivo_excel_incidencias

    # Inicializamos los primeros archivos
    archivo_csv_comas, archivo_csv_tabs, archivo_excel_buenos = generar_nombres_archivos_buenos(num_archivo_bueno)
    archivo_csv_incidencias, archivo_excel_incidencias = generar_nombres_archivos_incidencias(num_archivo_incidencias)

    try:
        with open(archivo_csv_comas, 'w', newline='', encoding='utf-8') as f_comas, \
             open(archivo_csv_tabs, 'w', newline='', encoding='utf-8') as f_tabs, \
             open(archivo_csv_incidencias, 'w', newline='', encoding='utf-8') as f_incidencias:
            
            writer_comas = csv.DictWriter(f_comas, fieldnames=fieldnames, delimiter=',')
            writer_tabs = csv.DictWriter(f_tabs, fieldnames=fieldnames, delimiter='\t')
            writer_incidencias_csv = csv.DictWriter(f_incidencias, fieldnames=fieldnames, delimiter=',')
            
            writer_comas.writeheader()
            writer_tabs.writeheader()
            writer_incidencias_csv.writeheader()

            wb_buenos = openpyxl.Workbook()
            ws_buenos = wb_buenos.active
            ws_buenos.append(fieldnames)
            
            wb_incidencias = openpyxl.Workbook()
            ws_incidencias = wb_incidencias.active
            ws_incidencias.append(fieldnames)

            for archivo in archivos_xml:
                consola.insert(tk.END, f"Procesando archivo: {archivo}\n")
                consola.update_idletasks()

                try:
                    for event, elem in ET.iterparse(archivo, events=("end",)):
                        if elem.tag == "DatosEnergeticosDelEdificio":
                            provincia = elem.find(".//Provincia").text or "N/A"
                            codigo_postal = elem.find(".//CodigoPostal").text or "N/A"
                            
                            # Normalizar provincia y código postal sin cambiar la provincia
                            provincia_normalizada = provincia.strip()  # No se cambia la provincia
                            codigo_postal_normalizado = normalizar_codigo_postal(codigo_postal)

                            registro = {
                                'Referencia Catastral': elem.find(".//ReferenciaCatastral").text or "N/A",
                                'Dirección': elem.find(".//Direccion").text or "N/A",
                                'Provincia': provincia_normalizada,
                                'Municipio': elem.find(".//Municipio").text or "N/A",
                                'Código Postal': codigo_postal_normalizado,
                                'Año Construcción': elem.find(".//AnoConstruccion").text or "N/A",
                                'Superficie Habitable': elem.find(".//SuperficieHabitable").text or "N/A",
                                'Demanda ACS': elem.find(".//DemandaDiariaACS").text or "N/A",
                                'Emisiones CO2 Global': elem.find(".//EmisionesCO2/Global").text or "N/A",
                                'Calificación Global': elem.find(".//Calificacion/EnergiaPrimariaNoRenovable/Global").text or "N/A",
                                'Fecha Certificación': elem.find(".//DatosDelCertificador/Fecha").text or "N/A",
                            }

                            # Validar si el código postal coincide con la provincia
                            if not validar_codigo_postal_provincia(provincia_normalizada, codigo_postal_normalizado):
                                # Si el código postal no coincide con la provincia, es una incidencia
                                guardar_incidencia_csv_excel(writer_incidencias_csv, ws_incidencias, registro, consola, archivo_csv_incidencias, archivo_excel_incidencias, row_incidencias)
                                row_incidencias += 1
                                total_incidencias += 1
                                registros_incidencias_por_archivo += 1

                                # Si se alcanzan los 30,000 registros de incidencias
                                if registros_incidencias_por_archivo >= limite_registros:
                                    consola.insert(tk.END, f"Guardando archivo de incidencias {num_archivo_incidencias}...\n")
                                    wb_incidencias.save(archivo_excel_incidencias)
                                    num_archivo_incidencias += 1
                                    archivo_csv_incidencias, archivo_excel_incidencias = generar_nombres_archivos_incidencias(num_archivo_incidencias)
                                    wb_incidencias = openpyxl.Workbook()
                                    ws_incidencias = wb_incidencias.active
                                    ws_incidencias.append(fieldnames)
                                    registros_incidencias_por_archivo = 0
                            else:
                                # Guardar como registro válido
                                guardar_registro_csv(writer_comas, registro, consola, archivo_csv_comas)
                                guardar_registro_csv(writer_tabs, registro, consola, archivo_csv_tabs)
                                guardar_registro_excel(ws_buenos, registro, registros_buenos_por_archivo + 1, consola, archivo_excel_buenos)
                                total_registros += 1
                                registros_buenos_por_archivo += 1

                                # Si se alcanzan los 30,000 registros buenos
                                if registros_buenos_por_archivo >= limite_registros:
                                    consola.insert(tk.END, f"Guardando archivo de registros buenos {num_archivo_bueno}...\n")
                                    wb_buenos.save(archivo_excel_buenos)
                                    num_archivo_bueno += 1
                                    archivo_csv_comas, archivo_csv_tabs, archivo_excel_buenos = generar_nombres_archivos_buenos(num_archivo_bueno)
                                    wb_buenos = openpyxl.Workbook()
                                    ws_buenos = wb_buenos.active
                                    ws_buenos.append(fieldnames)
                                    registros_buenos_por_archivo = 0

                            elem.clear()

                    consola.insert(tk.END, f"Total registros procesados en {archivo}: {total_registros}\n")
                    consola.see(tk.END)

                except Exception as e:
                    consola.insert(tk.END, f"Error procesando {archivo}: {e}\n")
                    consola.see(tk.END)

            # Guardar los archivos finales si no alcanzaron el límite de 30,000 registros
            wb_buenos.save(archivo_excel_buenos)
            wb_incidencias.save(archivo_excel_incidencias)
            consola.insert(tk.END, f"Archivo Excel de registros buenos guardado en: {archivo_excel_buenos}\n")
            consola.insert(tk.END, f"Archivo Excel de incidencias guardado en: {archivo_excel_incidencias}\n")

    except Exception as e:
        consola.insert(tk.END, f"Error al abrir los archivos CSV: {e}\n")
        consola.see(tk.END)

    consola.insert(tk.END, f"Indexación completada. Total de registros buenos: {total_registros}, total de incidencias: {total_incidencias}\n")
    consola.see(tk.END)



# Función para seleccionar archivos XML
def seleccionar_archivos_xml():
    archivos_xml = filedialog.askopenfilenames(title="Seleccionar archivos XML", filetypes=[("Archivos XML", "*.xml")])
    return archivos_xml

# Función que se ejecuta al presionar el botón para iniciar la indexación y la generación del CSV y Excel
def iniciar_indexacion():
    archivos_xml = seleccionar_archivos_xml()
    if not archivos_xml:
        consola.insert(tk.END, "Proceso cancelado por el usuario.\n")
        return

    archivo_csv_base = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Archivo CSV", "*.csv")], title="Guardar como CSV")
    if not archivo_csv_base:
        consola.insert(tk.END, "Proceso cancelado por el usuario.\n")
        return

    archivo_csv_comas = archivo_csv_base.replace('.csv', '_comas.csv')
    archivo_csv_tabs = archivo_csv_base.replace('.csv', '_tabulaciones.csv')
    archivo_excel = archivo_csv_base.replace('.csv', '.xlsx')

    consola.delete(1.0, tk.END)
    consola.insert(tk.END, f"Buscando archivos XML...\n")
    consola.update_idletasks()

    # Ejecutar la indexación y creación de los CSV y Excel en un hilo separado
    threading.Thread(target=indexar_archivos_xml_y_guardar, args=(archivos_xml, archivo_csv_comas, archivo_csv_tabs, archivo_excel, consola)).start()

# Crear la ventana principal
root = tk.Tk()
root.title("Indexador de Archivos XML a CSV y Excel")
root.geometry("600x400")

# Crear un área de texto desplazable para mostrar la consola
consola = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=70, height=20)
consola.pack(pady=10)

# Crear un botón para iniciar la indexación y guardar en CSV y Excel
boton_iniciar = tk.Button(root, text="Iniciar Indexación y Guardar CSV/Excel", command=iniciar_indexacion)
boton_iniciar.pack(pady=10)

# Mantener la ventana abierta
root.mainloop()
